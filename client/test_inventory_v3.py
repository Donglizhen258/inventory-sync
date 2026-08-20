from __future__ import annotations

import json
import multiprocessing
import os
import shutil
import tempfile
import threading
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from inventory_core import (
    APPENDED_LOCATION_HEADERS,
    CHANGE_TYPES,
    HEADERS,
    LEDGER_HEADERS,
    LEDGER_SHEET_NAME,
    LOCATIONS,
    LOCATION_HEADER,
    LEGACY_LOCATION_HEADERS,
    MAIN_COLUMN_WIDTHS,
    NEGATIVE_TYPES,
    PRE_LOCATION_HEADERS,
    PREVIOUS_HEADERS,
    POSITIVE_TYPES,
    PROJECTS,
    REQUESTERS,
    SHEET_NAME,
    InventoryRecord,
    current_balance,
    format_number,
    parse_datetime,
    search_materials,
    upgrade_workbook_schema,
    validate_record,
    workbook_suggestions,
    write_record,
)
from inventory_web import AppState, InventoryServer, RequestHandler


ROOT = Path(__file__).resolve().parent
TEMPLATE = ROOT / "assets" / "实验室库存管理模板.xlsx"
TEMP_V2 = Path(
    os.environ.get(
        "LAB_INVENTORY_TEMP_V2",
        ROOT / "test-data" / "库存管理TempV2.xlsx",
    )
)


def record(
    *,
    material: str = "测试电阻",
    specification: str = "10kΩ 1% 0603",
    requester: str = "王五",
    project: str = "项目A",
    quantity: float = 10,
    change_type: str = "入库",
    material_code: str = "",
    location: str = "",
    changed_at: datetime | None = None,
) -> InventoryRecord:
    return InventoryRecord(
        material=material,
        specification=specification,
        requester=requester,
        project=project,
        po="PO-TEST",
        changed_at=changed_at or datetime(2026, 7, 29, 10, 0),
        quantity=quantity,
        change_type=change_type,
        note="自动化测试",
        material_code=material_code,
        location=location,
    )


def request_json(url: str, payload: dict | None = None) -> dict:
    data = None
    method = "GET"
    headers = {}
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        method = "POST"
        headers["Content-Type"] = "application/json"
    request = urllib.request.Request(url, data=data, method=method, headers=headers)
    with urllib.request.urlopen(request, timeout=5) as response:
        assert response.status == 200
        return json.loads(response.read().decode("utf-8"))


def create_empty_test_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目A"
    sheet.append(HEADERS)
    sheet.append([None] * len(HEADERS))
    table = Table(displayName="InventoryTable", ref="A1:K2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(path)
    workbook.close()


def create_pre_location_test_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存管理"
    sheet.append(PRE_LOCATION_HEADERS)
    sheet.append(
        [
            "MAT-000001",
            "旧版连接器",
            "LEGACY-001",
            "项目A",
            "王五、孙八",
            "PO-OLD",
            datetime(2026, 7, 1, 9, 0),
            8,
            "",
            "旧数据备注",
        ]
    )
    table = Table(displayName="InventoryTable", ref="A1:J2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(path)
    workbook.close()


def create_appended_location_test_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(APPENDED_LOCATION_HEADERS)
    sheet.append(
        [
            "MAT-000777",
            "Legacy material",
            "LEGACY-777",
            "项目A",
            "王五",
            "PO-777",
            datetime(2026, 7, 1, 9, 0),
            8,
            "legacy detail",
            "legacy note",
            "\u8d27\u67b6H4",
        ]
    )
    table = Table(displayName="InventoryTable", ref="A1:K2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(path)
    workbook.close()


def create_legacy_location_test_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(LEGACY_LOCATION_HEADERS)
    sheet.append(
        [
            "MAT-000888",
            "Legacy stock order",
            "LEGACY-888",
            "项目B",
            "王五",
            "PO-888",
            datetime(2026, 7, 1, 9, 30),
            12,
            "货架A2",
            "legacy detail",
            "legacy note",
        ]
    )
    table = Table(displayName="InventoryTable", ref="A1:K2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(path)
    workbook.close()


def create_previous_layout_test_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(PREVIOUS_HEADERS)
    sheet.append(
        [
            "MAT-000999",
            "Previous layout material",
            "PREVIOUS-999",
            21,
            "项目C",
            "王五、孙八",
            "PO-999",
            datetime(2026, 8, 10, 13, 45),
            "货架H4",
            "previous detail",
            "previous note",
        ]
    )
    table = Table(displayName="InventoryTable", ref="A1:K2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(path)
    workbook.close()


def assert_table_layout(workbook) -> None:
    from inventory_core import is_project_sheet_title

    for sheet in workbook.worksheets:
        for table in sheet.tables.values():
            _, header_row, _, _ = range_boundaries(table.ref)
            for row in sheet[table.ref]:
                for cell in row:
                    is_main_detail = (
                        is_project_sheet_title(sheet.title)
                        and cell.column == HEADERS.index("变更明细") + 1
                        and cell.row > header_row
                    )
                    assert cell.alignment.horizontal == (
                        "left" if is_main_detail else "center"
                    )
                    assert cell.alignment.vertical == "center"
                    if is_main_detail:
                        assert cell.alignment.wrap_text is True
    for sheet in workbook.worksheets:
        if not is_project_sheet_title(sheet.title):
            continue
        for column, header in enumerate(HEADERS, start=1):
            column_letter = sheet.cell(1, column).column_letter
            assert abs(
                sheet.column_dimensions[column_letter].width
                - MAIN_COLUMN_WIDTHS[header]
            ) < 0.01


def assert_stock_conditional_formatting(sheet, last_row: int) -> None:
    groups = [
        (str(conditional_format.sqref), rules)
        for conditional_format, rules in sheet.conditional_formatting._cf_rules.items()
    ]
    assert len(groups) == 1
    assert groups[0][0] == ("D2" if last_row == 2 else f"D2:D{last_row}")
    assert {
        (rule.type, rule.operator, tuple(str(value) for value in rule.formula))
        for rule in groups[0][1]
    } == {
        ("cellIs", "equal", ("0",)),
        ("cellIs", "greaterThan", ("0",)),
    }


def assert_previous_layout_upgrade(workbook_path: Path) -> None:
    create_previous_layout_test_workbook(workbook_path)
    assert upgrade_workbook_schema(workbook_path) is True
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目C"]
        assert [main.cell(1, column).value for column in range(1, 12)] == HEADERS
        assert main.cell(2, 4).value == 21
        assert main.cell(2, 5).value == "货架H4"
        assert main.cell(2, 6).value == "项目C"
        assert main.cell(2, 7).value == "王五"
        assert main.cell(2, 8).value == "PO-999"
        assert main.cell(2, 9).value == datetime(2026, 8, 10)
        assert main.cell(2, 9).number_format == "yyyy-mm-dd"
        assert main.cell(2, 10).value == "previous detail"
        assert main.cell(2, 11).value == "previous note"
        assert_table_layout(workbook)
        assert_stock_conditional_formatting(main, 2)
    finally:
        workbook.close()
    assert upgrade_workbook_schema(workbook_path) is False


def assert_legacy_location_upgrade(workbook_path: Path) -> None:
    create_legacy_location_test_workbook(workbook_path)
    assert upgrade_workbook_schema(workbook_path) is True
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目B"]
        assert [
            main.cell(1, column).value
            for column in range(1, len(HEADERS) + 1)
        ] == HEADERS
        assert main.cell(2, 4).value == 12
        assert main.cell(2, 5).value == "货架A2"
        assert main.cell(2, 6).value == "项目B"
        assert main.cell(2, 9).value == datetime(2026, 7, 1)
        assert main.cell(2, 9).number_format == "yyyy-mm-dd"
        assert main.cell(2, 10).value == "legacy detail"
        assert main.cell(2, 11).value == "legacy note"
    finally:
        workbook.close()


def assert_appended_location_upgrade(workbook_path: Path) -> None:
    create_appended_location_test_workbook(workbook_path)
    assert upgrade_workbook_schema(workbook_path) is True

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目A"]
        assert [
            main.cell(1, column).value
            for column in range(1, len(HEADERS) + 1)
        ] == HEADERS
        assert main.cell(2, 5).value == "\u8d27\u67b6H4"
        assert main.cell(2, 10).value == "legacy detail"
        assert main.cell(2, 11).value == "legacy note"
        assert main.tables["InventoryTable"].ref == "A1:K2"
        assert len(main.tables["InventoryTable"].tableColumns) == len(HEADERS)
    finally:
        workbook.close()


def assert_pre_location_upgrade(workbook_path: Path) -> None:
    create_pre_location_test_workbook(workbook_path)
    assert upgrade_workbook_schema(workbook_path) is True
    backups = list(
        (workbook_path.parent / "backups").glob(
            f"{workbook_path.stem}_*.xlsx"
        )
    )
    assert len(backups) == 1

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目A"]
        assert [
            main.cell(1, column).value
            for column in range(1, len(HEADERS) + 1)
        ] == HEADERS
        assert main.cell(2, 5).value in (None, "")
        assert main.cell(2, 7).value == "王五"
        assert main.cell(2, 11).value == "旧数据备注"
        assert main.tables["InventoryTable"].ref == "A1:K2"
        assert len(main.tables["InventoryTable"].tableColumns) == len(HEADERS)
    finally:
        workbook.close()

    assert upgrade_workbook_schema(workbook_path) is False
    assert len(
        list(
            (workbook_path.parent / "backups").glob(
                f"{workbook_path.stem}_*.xlsx"
            )
        )
    ) == 1


def assert_manual_location_preserved(workbook_path: Path) -> None:
    create_empty_test_workbook(workbook_path)
    first = record(material="手工位置物料", specification="LOC-001")
    _, _, row = write_record(workbook_path, first)

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        workbook["项目A"].cell(row, 5).value = "手工填写区-Z"
        workbook.save(workbook_path)
    finally:
        workbook.close()

    matches = search_materials(workbook_path, "手工填写区-Z")
    assert len(matches) == 1
    assert matches[0]["location"] == "手工填写区-Z"

    blank_update = record(
        material="手工位置物料",
        specification="LOC-001",
        material_code=first.material_code,
        quantity=1,
        location="",
    )
    write_record(workbook_path, blank_update)
    assert blank_update.location == "手工填写区-Z"

    custom_update = record(
        material="手工位置物料",
        specification="LOC-001",
        material_code=first.material_code,
        quantity=1,
        location="试验台旁边",
    )
    write_record(workbook_path, custom_update)
    assert search_materials(workbook_path, "LOC-001")[0]["location"] == "试验台旁边"


def assert_preloaded_asset() -> None:
    from inventory_core import list_project_names, validate_workbook

    validate_workbook(TEMPLATE)
    workbook = load_workbook(TEMPLATE, data_only=False)
    try:
        project_names = list_project_names(TEMPLATE)
        assert project_names == ["项目B", "项目C", "项目A", "项目D"], project_names
        rows = []
        for name in project_names:
            main = workbook[name]
            assert [
                main.cell(1, column).value
                for column in range(1, len(HEADERS) + 1)
            ] == HEADERS
            sheet_rows = list(main.iter_rows(min_row=2, values_only=True))
            rows.extend(sheet_rows)
            assert len({row[0] for row in sheet_rows}) == len(sheet_rows)
            assert all(
                main.cell(row, 4).number_format == "#,##0;[Red]-#,##0"
                for row in range(2, main.max_row + 1)
            )
        assert len(rows) == 144
        assert len({row[0] for row in rows}) == 144
        assert rows[0][0] == "MAT-000001"
        assert rows[-1][0] == "MAT-000144"
        assert Counter(row[5] for row in rows) == {
            "项目B": 62,
            "项目C": 39,
            "项目A": 40,
            "项目D": 3,
        }
        assert sum(float(row[3] or 0) > 0 for row in rows) == 73
        assert sum(float(row[3] or 0) == 0 for row in rows) == 71
        assert all(float(row[3] or 0) >= 0 for row in rows)
        assert all(
            not row[7] or isinstance(row[7], str)
            for row in rows
        )
        assert all(
            not isinstance(row[8], datetime) or row[8].hour == 0
            for row in rows
        )
        assert all("系统期初导入" not in str(row[9] or "") for row in rows)
        assert all(row[4] in (None, "") for row in rows)
        for name in project_names:
            main = workbook[name]
            assert_stock_conditional_formatting(main, main.max_row)

        ledger = workbook[LEDGER_SHEET_NAME]
        assert ledger.max_row == 153
        assert [ledger.cell(1, column).value for column in range(1, 14)] == LEDGER_HEADERS
        assert all(ledger.cell(row, 9).value == "期初导入" for row in range(2, 154))
        assert all(
            ledger.cell(row, column).number_format == "#,##0;[Red]-#,##0"
            for row in range(2, 154)
            for column in (10, 11, 12)
        )
        assert ledger.tables["InventoryLedgerTable"].ref == "A1:M153"

        pending = workbook["待确认物料"]
        assert pending.max_row == 148
        assert pending.tables["PendingMaterialTable"].ref == "A1:M148"
        assert_table_layout(workbook)
        for name in project_names:
            main = workbook[name]
            assert all(
                main.cell(row, 7).value
                != "王五、孙八"
                for row in range(2, main.max_row + 1)
            )
    finally:
        workbook.close()

    result = search_materials(TEMPLATE, "EXAMPLE-MODEL", limit=5)
    assert result == [
        {
            "materialCode": "MAT-000002",
            "material": "示例物料 / 电源模块",
            "specification": "EXAMPLE-MODEL-15WR3",
            "project": "项目B",
            "requester": "王五",
            "po": "PO123456789",
            "location": "",
            "balance": 26.0,
        }
    ]
    assert workbook_suggestions(TEMPLATE)["projects"] == [
        "项目B",
        "项目C",
        "项目A",
        "项目D",
    ]
    assert workbook_suggestions(TEMPLATE)["locations"] == []
    assert len(LOCATIONS) == 32
    assert LOCATIONS[:4] == ["货架A1", "货架A2", "货架A3", "货架A4"]
    assert "货架F4" in LOCATIONS
    assert "货架F5" not in LOCATIONS
    assert LOCATIONS[-1] == "货架H4"
    assert CHANGE_TYPES == ["入库", "领用", "寄出", "报废", "退回", "整理"]
    assert POSITIVE_TYPES == {"入库", "退回", "整理"}
    assert NEGATIVE_TYPES == {"领用", "寄出", "报废"}


def assert_new_template_flow(workbook_path: Path) -> None:
    first = record(location="货架A1")
    before, after, first_row = write_record(workbook_path, first)
    assert before == 0
    assert after == 10
    assert first.material_code == "MAT-000001"

    second = record(
        requester="李四",
        project="项目B",
        quantity=-3,
        change_type="领用",
        material_code=first.material_code,
        location="货架B2",
        changed_at=datetime(2026, 7, 29, 11, 30),
    )
    before, after, second_row = write_record(workbook_path, second)
    assert before == 10
    assert after == 7
    assert second_row == first_row
    assert current_balance(workbook_path, material_code=first.material_code) == 7

    try:
        write_record(
            workbook_path,
            record(
                material="错误物料名称",
                specification="错误规格",
                material_code=first.material_code,
                quantity=1,
            ),
        )
    except ValueError as exc:
        assert "名称或规格已发生变化" in str(exc)
    else:
        raise AssertionError("编码与名称/规格不一致时仍被写入")

    variant = record(specification="10kΩ 1% 0805", quantity=5)
    write_record(workbook_path, variant)
    assert variant.material_code == "MAT-000002"
    assert variant.material_code != first.material_code

    matches = search_materials(workbook_path, "测试 0603")
    assert len(matches) == 1
    assert matches[0]["materialCode"] == first.material_code
    assert matches[0]["balance"] == 7
    assert matches[0]["project"] == "项目A"
    assert matches[0]["requester"] == "李四"
    assert matches[0]["location"] == "货架B2"

    try:
        write_record(
            workbook_path,
            record(
                quantity=-100,
                change_type="领用",
                material_code=first.material_code,
            ),
        )
    except ValueError as exc:
        assert "库存不足" in str(exc)
    else:
        raise AssertionError("负库存操作未被拦截")

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目A"]
        assert [
            main.cell(1, col).value
            for col in range(1, len(HEADERS) + 1)
        ] == HEADERS
        assert main.cell(first_row, 4).value == 7
        assert main.cell(first_row, 5).value == "货架B2"
        detail = str(main.cell(first_row, 10).value)
        assert detail == (
            "2026-7-29：王五入库10个；\n"
            "2026-7-29：李四领用3个；"
        )
        assert main.cell(first_row, 10).alignment.wrap_text is True
        assert main.cell(first_row, 4).number_format == "#,##0;[Red]-#,##0"
        assert main.cell(first_row, 9).value == datetime(2026, 7, 29)
        assert main.cell(first_row, 9).number_format == "yyyy-mm-dd"
        assert main.row_dimensions[first_row].height >= 36
        assert main.tables["InventoryTable"].ref == "A1:K3"

        ledger = workbook[LEDGER_SHEET_NAME]
        assert ledger.sheet_state == "hidden"
        assert [ledger.cell(1, col).value for col in range(1, 14)] == LEDGER_HEADERS
        assert ledger.cell(2, 1).value == "TXN-00000001"
        assert ledger.cell(3, 1).value == "TXN-00000002"
        assert ledger.cell(4, 1).value == "TXN-00000003"
        assert ledger.cell(2, 10).number_format == "#,##0;[Red]-#,##0"
        assert ledger.cell(2, 11).number_format == "#,##0;[Red]-#,##0"
        assert ledger.cell(2, 12).number_format == "#,##0;[Red]-#,##0"
        assert ledger.cell(2, 8).value == datetime(2026, 7, 29)
        assert ledger.cell(3, 8).value == datetime(2026, 7, 29)
        assert ledger.cell(2, 8).number_format == "yyyy-mm-dd"
        assert ledger.tables["InventoryLedgerTable"].ref == "A1:M4"
        assert_table_layout(workbook)
        assert_stock_conditional_formatting(main, 3)
    finally:
        workbook.close()

    try:
        validate_record(
            "测试",
            "王五",
            "项目A",
            "入库",
            "NaN",
        )
    except ValueError as exc:
        assert "有限数字" in str(exc)
    else:
        raise AssertionError("NaN 未被拦截")

    for invalid_quantity, expected_message in (
        ("0.004", "必须是整数"),
        ("1.5", "必须是整数"),
        ("1000000000001", "变动数量过大"),
    ):
        try:
            validate_record(
                "测试",
                "王五",
                "项目A",
                "入库",
                invalid_quantity,
            )
        except ValueError as exc:
            assert expected_message in str(exc)
        else:
            raise AssertionError(f"非法数量 {invalid_quantity} 未被拦截")

    for change_type in POSITIVE_TYPES:
        assert validate_record(
            "测试",
            "王五",
            "项目A",
            change_type,
            "1",
        ) == 1
        try:
            validate_record(
                "测试",
                "王五",
                "项目A",
                change_type,
                "-1",
            )
        except ValueError as exc:
            assert "应填写正数" in str(exc)
        else:
            raise AssertionError(f"{change_type}错误接受了负数")

    for change_type in NEGATIVE_TYPES:
        assert validate_record(
            "测试",
            "王五",
            "项目A",
            change_type,
            "-1",
        ) == -1
        try:
            validate_record(
                "测试",
                "王五",
                "项目A",
                change_type,
                "1",
            )
        except ValueError as exc:
            assert "应填写负数" in str(exc)
        else:
            raise AssertionError(f"{change_type}错误接受了正数")

    try:
        validate_record(
            "测试",
            "王五",
            "项目A",
            "盘盈",
            "1",
        )
    except ValueError as exc:
        assert "有效的库存管理类型" in str(exc)
    else:
        raise AssertionError("旧库存管理类型仍被接受")

    try:
        validate_record(
            "=HYPERLINK(\"https://example.invalid\")",
            "王五",
            "项目A",
            "入库",
            "1",
        )
    except ValueError as exc:
        assert "不能以等号开头" in str(exc)
    else:
        raise AssertionError("公式型文本未被拦截")


def assert_temp_v2_upgrade(workbook_path: Path) -> None:
    update = record(
        material="可乐",
        specification="",
        quantity=-2,
        change_type="领用",
        location="货架C3",
    )
    before, after, _ = write_record(workbook_path, update)
    assert before == 20
    assert after == 18
    assert update.material_code == "MAT-000001"

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["Sheet1"]
        assert [
            main.cell(1, col).value
            for col in range(1, len(HEADERS) + 1)
        ] == HEADERS
        assert main.cell(2, 1).value == "MAT-000001"
        assert main.cell(2, 4).value == 18
        assert main.cell(2, 5).value == "货架C3"
        assert len(main.tables) == 1
        table = next(iter(main.tables.values()))
        assert table.ref == "A1:K2"
        assert len(table.tableColumns) == len(HEADERS)
        assert table.autoFilter is None or table.autoFilter.ref == "A1:K2"
        assert workbook[LEDGER_SHEET_NAME].sheet_state == "hidden"
        assert len(workbook[LEDGER_SHEET_NAME].tables) == 1
    finally:
        workbook.close()


def assert_identity_normalization(workbook_path: Path) -> None:
    first = record(
        material="精密电阻　Ａ",
        specification="１０ＫΩ　０６０３",
        quantity=2,
    )
    _, _, first_row = write_record(workbook_path, first)
    second = record(
        material="精密电阻 A",
        specification="10KΩ 0603",
        quantity=3,
    )
    _, after, second_row = write_record(workbook_path, second)
    assert second.material_code == first.material_code
    assert second_row == first_row
    assert after == 5

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目A"]
        main.cell(first_row, 4).value = "=2+3"
        workbook.save(workbook_path)
    finally:
        workbook.close()
    try:
        current_balance(workbook_path, material_code=first.material_code)
    except ValueError as exc:
        assert "不能使用 Excel 公式" in str(exc)
    else:
        raise AssertionError("公式库存未被拦截")


def assert_workbook_suggestion_flow(workbook_path: Path) -> None:
    imported = record(
        material="历史连接器",
        specification="HIST-001",
        requester="历史需求者",
        project="Legacy Project",
        quantity=4,
        location="自定义柜-3",
    )
    write_record(workbook_path, imported)
    normalized_duplicate = record(
        material="全角项目测试",
        specification="FULLWIDTH-001",
        requester="  王五  ",
        project="ＭＶＳＳＴ",
        quantity=1,
        location="货架A1",
    )
    write_record(workbook_path, normalized_duplicate)

    workbook_values = workbook_suggestions(workbook_path)
    assert workbook_values["requesters"] == [
        "历史需求者",
        "王五",
    ]
    assert workbook_values["projects"] == [
        "项目A",
        "Legacy Project",
        "ＭＶＳＳＴ",
    ]
    assert workbook_values["locations"] == ["自定义柜-3", "货架A1"]

    state = AppState.__new__(AppState)
    state.workbook_path = workbook_path
    state.lock = threading.Lock()
    payload = state.state_payload()
    assert payload["requesters"] == REQUESTERS
    assert payload["projects"] == [
        "项目A",
        "Legacy Project",
        "ＭＶＳＳＴ",
    ]
    assert payload["locations"] == [*LOCATIONS, "自定义柜-3"]


def assert_frontend_material_candidate_contract() -> None:
    source = (ROOT / "web" / "app.js").read_text(encoding="utf-8")
    html = (ROOT / "web" / "index.html").read_text(encoding="utf-8")
    assert 'const DEFAULT_PROJECTS = ["项目A", "项目B", "项目C", "项目D"];' in source
    assert 'const SHELF_LETTERS = ["A", "B", "C", "D", "E", "F", "G", "H"];' in source
    assert "F: 4" in source
    assert "DEFAULT_LOCATIONS = SHELF_LETTERS.flatMap" in source
    assert '"张三"' in source
    assert '"王小明"' in source
    assert '"大力"' not in source
    assert '"智杭"' not in source
    assert "<title>ERL库存管理</title>" in html
    assert "<h1>ERL库存管理</h1>" in html
    assert "ERL INVENTORY" in html
    assert "入库、退回、整理填正数；领用、寄出、报废填负数。" in html
    assert 'id="changedAt" name="changedAt" type="date"' in html
    assert '<button id="nowButton" type="button">今天</button>' in html
    assert "function localDate()" in source
    assert "localDateTime" not in source
    assert 'candidate.project || "—"' in source
    assert 'elements.requester.value = candidate.requester || "";' in source
    assert 'elements.project.value = candidate.project || "";' in source
    assert 'elements.po.value = candidate.po || "";' in source
    assert 'elements.location.value = candidate.location || "";' in source
    assert 'id="locationSuggestions"' in html
    assert 'id="shelfLetter"' in html
    assert 'id="shelfNumber"' in html
    assert "自动拼接，也可手动输入" in html
    assert "货架 A–H 均可选 1–4" in html
    assert "syncLocationFromShelf" in source
    assert "<th>物料放置位置</th>" in html
    assert "maximumFractionDigits: 0" in source
    assert 'placeholder.textContent = "请选择";' in source
    assert "if (!input.value && suggestions.length)" not in source
    assert 'elements.clearButton.addEventListener("click", () => clearForm());' in source
    assert 'id="clearButton"' in html
    assert 'step="1"' in html
    assert format_number(1234.4) == "1,234"


def concurrent_worker(
    workbook_path: str,
    material: str,
    start_event,
    result_queue,
) -> None:
    try:
        start_event.wait(10)
        item = record(
            material=material,
            specification="并发规格",
            requester="并发用户",
            project="并发项目",
            quantity=1,
        )
        write_record(Path(workbook_path), item)
        result_queue.put(("ok", item.material_code))
    except Exception as exc:
        result_queue.put(("error", repr(exc)))


def assert_cross_process_lock(workbook_path: Path) -> None:
    context = multiprocessing.get_context("spawn")
    start_event = context.Event()
    result_queue = context.Queue()
    processes = [
        context.Process(
            target=concurrent_worker,
            args=(str(workbook_path), f"并发物料{index}", start_event, result_queue),
        )
        for index in (1, 2)
    ]
    for process in processes:
        process.start()
    start_event.set()
    for process in processes:
        process.join(20)
        assert process.exitcode == 0

    results = [result_queue.get(timeout=5) for _ in processes]
    assert all(status == "ok" for status, _ in results), results
    assert {value for _, value in results} == {"MAT-000001", "MAT-000002"}

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["并发项目"]
        codes = {
            str(main.cell(row, 1).value)
            for row in range(2, main.max_row + 1)
            if main.cell(row, 1).value
        }
        assert codes == {"MAT-000001", "MAT-000002"}
        ledger = workbook[LEDGER_SHEET_NAME]
        assert sum(
            1
            for row in range(2, ledger.max_row + 1)
            if ledger.cell(row, 1).value
        ) == 2
    finally:
        workbook.close()


def assert_http_flow(workbook_path: Path) -> None:
    state = AppState.__new__(AppState)
    state.workbook_path = workbook_path
    state.lock = threading.Lock()
    state.sync_lock = threading.RLock()
    state.server_url = ""
    state.sync_token = ""
    state.logged_in = False
    state.online = False
    state.role = "user"
    state.baseline_version = None
    state.pending_push = False
    state.last_sync_at = None
    server = InventoryServer(("127.0.0.1", 0), RequestHandler, state)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{server.server_address[1]}"
    try:
        health = request_json(f"{base}/api/health")
        assert health["ok"] is True

        query = urllib.parse.urlencode({"q": "测试 0603", "limit": 12})
        materials = request_json(f"{base}/api/materials?{query}")
        assert materials["materials"][0]["materialCode"] == "MAT-000001"
        assert materials["materials"][0]["project"] == "项目A"

        recent = request_json(f"{base}/api/recent")
        assert recent["records"]
        assert recent["records"][0]["material_code"].startswith("MAT-")
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def assert_reorganize_flow(workbook_path: Path) -> None:
    create_empty_test_workbook(workbook_path)
    initial = record(
        material="整理测试物料",
        specification="ORG-001",
        quantity=5,
        changed_at=datetime(2026, 8, 11, 9, 30),
    )
    write_record(workbook_path, initial)
    adjustment = record(
        material="整理测试物料",
        specification="ORG-001",
        material_code=initial.material_code,
        quantity=2,
        change_type="整理",
        changed_at=datetime(2026, 8, 12, 18, 45),
    )
    before, after, row = write_record(workbook_path, adjustment)
    assert before == 5
    assert after == 7

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        main = workbook["项目A"]
        ledger = workbook[LEDGER_SHEET_NAME]
        assert main.cell(row, 4).value == 7
        assert str(main.cell(row, 10).value).endswith(
            "2026-8-12：王五整理2个；"
        )
        assert ledger.cell(3, 8).value == datetime(2026, 8, 12)
        assert ledger.cell(3, 9).value == "整理"
        assert ledger.cell(3, 10).value == 2
    finally:
        workbook.close()

    assert parse_datetime("2026/08/13") == datetime(2026, 8, 13)
    try:
        parse_datetime("2026-08-13 09:15")
    except ValueError as exc:
        assert "年-月-日" in str(exc)
    else:
        raise AssertionError("带小时分钟的变动时间未被拒绝")


def main() -> None:
    assert_preloaded_asset()
    with tempfile.TemporaryDirectory(prefix="inventory-v3-") as temp:
        temp_dir = Path(temp)
        workbook_path = temp_dir / "inventory.xlsx"
        create_empty_test_workbook(workbook_path)
        assert_new_template_flow(workbook_path)
        assert_http_flow(workbook_path)

        reorganize_path = temp_dir / "reorganize.xlsx"
        assert_reorganize_flow(reorganize_path)

        identity_path = temp_dir / "identity.xlsx"
        create_empty_test_workbook(identity_path)
        assert_identity_normalization(identity_path)

        suggestions_path = temp_dir / "suggestions.xlsx"
        create_empty_test_workbook(suggestions_path)
        assert_workbook_suggestion_flow(suggestions_path)
        assert_frontend_material_candidate_contract()

        legacy_path = temp_dir / "pre-location.xlsx"
        assert_pre_location_upgrade(legacy_path)

        appended_location_path = temp_dir / "appended-location.xlsx"
        assert_appended_location_upgrade(appended_location_path)

        legacy_location_path = temp_dir / "legacy-stock-order.xlsx"
        assert_legacy_location_upgrade(legacy_location_path)

        previous_layout_path = temp_dir / "previous-location-order.xlsx"
        assert_previous_layout_upgrade(previous_layout_path)

        manual_location_path = temp_dir / "manual-location.xlsx"
        assert_manual_location_preserved(manual_location_path)

        concurrent_path = temp_dir / "concurrent.xlsx"
        create_empty_test_workbook(concurrent_path)
        assert_cross_process_lock(concurrent_path)

        if TEMP_V2.is_file():
            temp_v2_path = temp_dir / "temp-v2.xlsx"
            shutil.copy2(TEMP_V2, temp_v2_path)
            assert_temp_v2_upgrade(temp_v2_path)

    print(json.dumps({"status": "ok"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
