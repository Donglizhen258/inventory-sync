from __future__ import annotations

import ctypes
import hashlib
import json
import math
import os
import re
import shutil
import sys
import threading
import unicodedata
import uuid
from contextlib import contextmanager
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_NAME = "库存管理"
PRE_LOCATION_HEADERS = [
    "物料编码",
    "材料名称",
    "规格型号",
    "项目",
    "需求者",
    "PO",
    "入库时间",
    "库存数量",
    "变更明细",
    "备注",
]
LOCATION_HEADER = "物料放置位置"
APPENDED_LOCATION_HEADERS = [*PRE_LOCATION_HEADERS, LOCATION_HEADER]
LEGACY_LOCATION_HEADERS = [
    *PRE_LOCATION_HEADERS[:8],
    LOCATION_HEADER,
    *PRE_LOCATION_HEADERS[8:],
]
PREVIOUS_HEADERS = [
    "物料编码",
    "材料名称",
    "规格型号",
    "库存数量",
    "项目",
    "需求者",
    "PO",
    "入库时间",
    LOCATION_HEADER,
    "变更明细",
    "备注",
]
HEADERS = [
    "物料编码",
    "材料名称",
    "规格型号",
    "库存数量",
    LOCATION_HEADER,
    "项目",
    "需求者",
    "PO",
    "入库时间",
    "变更明细",
    "备注",
]
MAIN_COLUMN_WIDTHS = {
    "物料编码": 15,
    "材料名称": 30,
    "规格型号": 28,
    "库存数量": 12,
    LOCATION_HEADER: 16,
    "项目": 12,
    "需求者": 26,
    "PO": 22,
    "入库时间": 14,
    "变更明细": 58,
    "备注": 40,
}
TEMP_V2_HEADERS = [
    "材料名称",
    "项目",
    "需求者",
    "PO",
    "入库时间",
    "库存数量",
    "变更明细",
    "备注",
]
OLD_HEADERS = [
    "材料名称",
    "需求者",
    "PO",
    "变动时间",
    "变动数量",
    "剩余数量",
    "库存管理",
    "项目",
    "备注",
]
LEDGER_SHEET_NAME = "_变更流水"
LEDGER_HEADERS = [
    "流水编号",
    "物料编码",
    "材料名称",
    "规格型号",
    "项目",
    "需求者",
    "PO",
    "变动时间",
    "库存管理",
    "变动数量",
    "变动前库存",
    "变动后库存",
    "备注",
]
REQUESTERS = [
    "张三",
    "李四",
    "王五",
    "赵六",
    "钱七",
    "孙八",
    "周九",
    "吴十",
    "郑十一",
    "王小明",
]
REQUESTER_REPLACEMENTS = {
    "王五、孙八": "王五",
}
PROJECTS = ["项目A", "项目B", "项目C", "项目D"]
SHELF_NUMBER_LIMITS = {
    letter: 4
    for letter in "ABCDEFGH"
}
LOCATIONS = [
    f"货架{letter}{number}"
    for letter, maximum in SHELF_NUMBER_LIMITS.items()
    for number in range(1, maximum + 1)
]
CHANGE_TYPES = ["入库", "领用", "寄出", "报废", "退回", "整理"]
POSITIVE_TYPES = {"入库", "退回", "整理"}
NEGATIVE_TYPES = {"领用", "寄出", "报废"}
MAX_BACKUPS = 20
WRITE_LOCK = threading.Lock()
MAX_ABS_QUANTITY = Decimal("1000000000000")
MATERIAL_CODE_PATTERN = re.compile(r"^MAT-(\d{6,})$", re.IGNORECASE)
LEDGER_CODE_PATTERN = re.compile(r"^TXN-(\d{8,})$", re.IGNORECASE)


@dataclass(frozen=True)
class SheetLayout:
    sheet: Any
    header_row: int
    columns: dict[str, int]
    temp_v2: bool = False
    pre_location: bool = False
    appended_location: bool = False
    legacy_stock_order: bool = False
    location_after_date: bool = False


@dataclass
class InventoryRecord:
    material: str
    requester: str
    po: str
    changed_at: datetime
    quantity: float
    change_type: str
    project: str
    note: str
    specification: str = ""
    material_code: str = ""
    location: str = ""


def application_dir() -> Path:
    override = os.environ.get("LAB_INVENTORY_APP_DIR")
    if override:
        return Path(override).resolve()
    # 数据存到当前用户的专属目录（系统保护区域），避免明文 Excel 留在程序文件夹
    base = (
        os.environ.get("LOCALAPPDATA")
        or os.environ.get("APPDATA")
        or str(Path.home())
    )
    return Path(base) / "ERL库存管理"


def bundled_template() -> Path:
    if getattr(sys, "frozen", False):
        bundle_root = Path(getattr(sys, "_MEIPASS"))
    else:
        bundle_root = Path(__file__).resolve().parent
    return bundle_root / "assets" / "实验室库存管理模板.xlsx"


APP_DIR = application_dir()
SETTINGS_PATH = APP_DIR / "settings.json"
DEFAULT_WORKBOOK = APP_DIR / "data" / "实验室库存管理.xlsx"


def load_settings() -> dict[str, Any]:
    if not SETTINGS_PATH.exists():
        return {}
    try:
        return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_settings(workbook_path: Path) -> None:
    APP_DIR.mkdir(parents=True, exist_ok=True)
    SETTINGS_PATH.write_text(
        json.dumps({"workbook_path": str(workbook_path)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _create_empty_workbook(path: Path) -> None:
    """无内置模板时自举：生成一个空的标准工作簿（默认项目 sheet + 隐藏流水 sheet）。"""
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    _create_project_sheet(workbook, "未分类")
    _ensure_ledger(workbook)
    workbook.save(path)
    workbook.close()


def ensure_default_workbook() -> Path:
    selected = load_settings().get("workbook_path")
    if selected and Path(selected).is_file():
        selected_path = Path(selected).resolve()
        upgrade_workbook_schema(selected_path)
        return selected_path

    DEFAULT_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    created = False
    if not DEFAULT_WORKBOOK.exists():
        source = bundled_template()
        if source.exists():
            shutil.copy2(source, DEFAULT_WORKBOOK)
        else:
            _create_empty_workbook(DEFAULT_WORKBOOK)
        created = True
    if created:
        _hide_existing_ledger(DEFAULT_WORKBOOK)
    upgrade_workbook_schema(DEFAULT_WORKBOOK)
    save_settings(DEFAULT_WORKBOOK)
    return DEFAULT_WORKBOOK


def _text(value: Any) -> str:
    return str("" if value is None else value).strip()


def normalized_material(value: Any) -> str:
    return _normalized_identity(value)


def normalized_specification(value: Any) -> str:
    return _normalized_identity(value)


def _normalized_identity(value: Any) -> str:
    normalized = unicodedata.normalize("NFKC", _text(value))
    return " ".join(normalized.split()).casefold()


def numeric_value(value: Any) -> float:
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _stock_value(value: Any) -> float:
    if isinstance(value, str) and value.lstrip().startswith("="):
        raise ValueError("库存数量必须是数值，不能使用 Excel 公式。")
    return numeric_value(value)


def _row_values(sheet: Any, row: int, count: int) -> list[str]:
    return [_text(sheet.cell(row, column).value) for column in range(1, count + 1)]


PENDING_SHEET_NAME = "待确认物料"
RESERVED_SHEET_NAMES = {LEDGER_SHEET_NAME, PENDING_SHEET_NAME}


def is_project_sheet_title(title: str) -> bool:
    """项目 sheet 命名规则：非空、不以 _ 开头、不是保留名。"""
    if not title or not title.strip():
        return False
    if title.startswith("_"):
        return False
    return title not in RESERVED_SHEET_NAMES


def list_project_sheets(workbook: Any) -> list[Any]:
    return [
        sheet
        for sheet in workbook.worksheets
        if is_project_sheet_title(sheet.title)
    ]


def list_project_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return [sheet.title for sheet in list_project_sheets(workbook)]
    finally:
        workbook.close()


def find_project_sheet(workbook: Any, project: str) -> Any | None:
    for sheet in list_project_sheets(workbook):
        if sheet.title == project:
            return sheet
    return None


def validate_project_name(project: str) -> str:
    project = str(project or "").strip()
    if not project:
        raise ValueError("项目名称不能为空。")
    if len(project) > 31:
        raise ValueError("项目名称不能超过 31 个字符。")
    if any(character in project for character in ":\\/?*[]"):
        raise ValueError("项目名称不能包含 : \\ / ? * [ ] 等字符。")
    if not is_project_sheet_title(project):
        raise ValueError("项目名称不能以 _ 开头，也不能与系统工作表重名。")
    return project


def _create_project_sheet(workbook: Any, project: str) -> Any:
    """新建一个标准项目 sheet（表头第一行 + 列宽 + 表格）。"""
    project = validate_project_name(project)
    existing = find_project_sheet(workbook, project)
    if existing is not None:
        return existing
    sheet = workbook.create_sheet(title=project)
    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(1, index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="2F75B5",
            end_color="2F75B5",
            fill_type="solid",
        )
    layout = SheetLayout(
        sheet,
        1,
        {name: index + 1 for index, name in enumerate(HEADERS)},
    )
    _apply_main_column_widths(layout)
    _apply_center_alignment(workbook, layout)
    _apply_date_only_formats(workbook, layout)
    _rebuild_main_table(workbook, sheet, header_row=1, last_row=1)
    _normalize_stock_conditional_formatting(layout)
    return sheet


def _locate_project_layout(workbook: Any, project: str) -> SheetLayout:
    sheet = find_project_sheet(workbook, project)
    if sheet is None:
        sheet = _create_project_sheet(workbook, project)
    return SheetLayout(
        sheet,
        1,
        {name: index + 1 for index, name in enumerate(HEADERS)},
    )


def _split_single_sheet_to_projects(workbook: Any) -> bool:
    """把旧版单主表（库存管理）按项目列拆分为多项目 sheet。返回是否执行了拆分。"""
    if SHEET_NAME not in workbook.sheetnames:
        return False
    try:
        layout = _locate_main_sheet(workbook)
    except ValueError:
        return False
    if layout.sheet.title != SHEET_NAME:
        return False
    if "项目" not in layout.columns:
        return False

    rows = []
    for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
        material = _text(layout.sheet.cell(row, layout.columns["材料名称"]).value)
        if not material:
            continue
        project = _text(layout.sheet.cell(row, layout.columns["项目"]).value)
        if not project:
            project = "未分类"
        values = {}
        for name in HEADERS:
            column = layout.columns.get(name)
            values[name] = (
                layout.sheet.cell(row, column).value if column else None
            )
        rows.append((project, values))

    # 先删除旧单表，确保后续创建的项目 sheet 使用统一表名 InventoryTable
    del workbook[SHEET_NAME]

    for project, values in rows:
        sheet = find_project_sheet(workbook, project)
        if sheet is None:
            sheet = _create_project_sheet(workbook, project)
        target_row = sheet.max_row + 1
        if target_row > 2:
            _copy_row_style(sheet, 2, target_row, len(HEADERS))
        for index, name in enumerate(HEADERS, start=1):
            cell = sheet.cell(target_row, index)
            cell.value = values.get(name)
            if name == "入库时间" and isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd"
            if name == "库存数量":
                cell.number_format = "#,##0;[Red]-#,##0"

    # 重新应用格式：非明细列水平居中，变更明细正文左对齐并自动换行
    for project in {project for project, _ in rows}:
        layout = _locate_project_layout(workbook, project)
        _apply_center_alignment(workbook, layout)
        detail_column = layout.columns["变更明细"]
        for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
            cell = layout.sheet.cell(row, detail_column)
            if (
                cell.alignment.horizontal == "left"
                and cell.alignment.wrap_text is True
            ):
                continue
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
    return True


def _locate_main_sheet(workbook: Any) -> SheetLayout:
    candidates = []
    preferred = {"库存管理": 0, "Sheet1": 1}
    for sheet in workbook.worksheets:
        if sheet.title == LEDGER_SHEET_NAME:
            continue
        candidates.append((preferred.get(sheet.title, 2), sheet))
    candidates.sort(key=lambda item: item[0])

    old_format_found = False
    for _, sheet in candidates:
        for row in range(1, min(sheet.max_row, 10) + 1):
            if _row_values(sheet, row, len(HEADERS)) == HEADERS:
                return SheetLayout(
                    sheet,
                    row,
                    {name: index + 1 for index, name in enumerate(HEADERS)},
                )
            if _row_values(sheet, row, len(PREVIOUS_HEADERS)) == PREVIOUS_HEADERS:
                return SheetLayout(
                    sheet,
                    row,
                    {
                        name: index + 1
                        for index, name in enumerate(PREVIOUS_HEADERS)
                    },
                    location_after_date=True,
                )
            if (
                _row_values(sheet, row, len(LEGACY_LOCATION_HEADERS))
                == LEGACY_LOCATION_HEADERS
            ):
                return SheetLayout(
                    sheet,
                    row,
                    {
                        name: index + 1
                        for index, name in enumerate(LEGACY_LOCATION_HEADERS)
                    },
                    legacy_stock_order=True,
                )
            if (
                _row_values(sheet, row, len(APPENDED_LOCATION_HEADERS))
                == APPENDED_LOCATION_HEADERS
            ):
                return SheetLayout(
                    sheet,
                    row,
                    {
                        name: index + 1
                        for index, name in enumerate(APPENDED_LOCATION_HEADERS)
                    },
                    appended_location=True,
                )
            if (
                _row_values(sheet, row, len(PRE_LOCATION_HEADERS))
                == PRE_LOCATION_HEADERS
            ):
                return SheetLayout(
                    sheet,
                    row,
                    {
                        name: index + 1
                        for index, name in enumerate(PRE_LOCATION_HEADERS)
                    },
                    pre_location=True,
                )
            if _row_values(sheet, row, len(TEMP_V2_HEADERS)) == TEMP_V2_HEADERS:
                return SheetLayout(
                    sheet,
                    row,
                    {name: index + 1 for index, name in enumerate(TEMP_V2_HEADERS)},
                    temp_v2=True,
                )
            if _row_values(sheet, row, len(OLD_HEADERS)) == OLD_HEADERS:
                old_format_found = True

    if old_format_found:
        raise ValueError(
            "所选 Excel 是旧版“每次变动一行”的九列格式，不能直接写入。"
            "请先迁移到新版“一物料一行”模板。"
        )
    raise ValueError(
        "未找到可用的库存主表。表头应为新版十一列格式、"
        "兼容的十列格式、位置在末列的过渡格式，"
        "或库存管理TempV2的八列格式。"
    )


def validate_workbook(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError("请选择 .xlsx 格式的 Excel 文件。")
    if not path.is_file():
        raise FileNotFoundError(f"找不到 Excel 文件：{path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        _locate_main_sheet(workbook)
    finally:
        workbook.close()


def _copy_row_style(sheet: Any, source_row: int, target_row: int, columns: int) -> None:
    if source_row < 1 or source_row == target_row:
        return
    for column in range(1, columns + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    if source_row in sheet.row_dimensions:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _extend_header_table(
    sheet: Any,
    *,
    header_row: int,
    last_row: int,
    last_column: int,
) -> None:
    for table in sheet.tables.values():
        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        if min_row != header_row or min_column != 1:
            continue
        table.ref = (
            f"A{header_row}:"
            f"{get_column_letter(max(max_column, last_column))}"
            f"{max(max_row, last_row)}"
        )
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref


def _next_table_name(workbook: Any, base_name: str) -> str:
    existing = {
        table.name
        for worksheet in workbook.worksheets
        for table in worksheet.tables.values()
    }
    if base_name not in existing:
        return base_name
    suffix = 2
    while f"{base_name}{suffix}" in existing:
        suffix += 1
    return f"{base_name}{suffix}"


def _rebuild_main_table(
    workbook: Any,
    sheet: Any,
    *,
    header_row: int,
    last_row: int,
) -> None:
    for table_name in list(sheet.tables.keys()):
        del sheet.tables[table_name]
    table = Table(
        displayName=_next_table_name(workbook, "InventoryTable"),
        ref=(
            f"A{header_row}:"
            f"{get_column_letter(len(HEADERS))}"
            f"{max(header_row + 1, last_row)}"
        ),
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _hide_existing_ledger(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    temp_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        if LEDGER_SHEET_NAME not in workbook.sheetnames:
            return
        ledger = workbook[LEDGER_SHEET_NAME]
        if ledger.sheet_state == "hidden":
            return
        ledger.sheet_state = "hidden"
        workbook.save(temp_path)
        workbook.close()
        os.replace(temp_path, path)
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


@contextmanager
def _cross_process_workbook_lock(path: Path):
    if sys.platform != "win32":
        yield
        return

    mutex_key = hashlib.sha256(
        os.path.normcase(str(path.resolve())).encode("utf-8")
    ).hexdigest()
    mutex_name = f"Local\\LabInventoryWorkbook-{mutex_key}"
    kernel32 = ctypes.windll.kernel32
    kernel32.CreateMutexW.argtypes = [
        ctypes.c_void_p,
        ctypes.c_bool,
        ctypes.c_wchar_p,
    ]
    kernel32.CreateMutexW.restype = ctypes.c_void_p
    kernel32.WaitForSingleObject.argtypes = [ctypes.c_void_p, ctypes.c_uint32]
    kernel32.WaitForSingleObject.restype = ctypes.c_uint32
    kernel32.ReleaseMutex.argtypes = [ctypes.c_void_p]
    kernel32.CloseHandle.argtypes = [ctypes.c_void_p]

    handle = kernel32.CreateMutexW(None, False, mutex_name)
    if not handle:
        raise OSError("无法创建库存文件写入锁。")

    wait_result = kernel32.WaitForSingleObject(handle, 15_000)
    wait_object_0 = 0x00000000
    wait_abandoned = 0x00000080
    try:
        if wait_result not in (wait_object_0, wait_abandoned):
            raise PermissionError("另一个库存程序正在写入该 Excel，请稍后重试。")
        yield
    finally:
        if wait_result in (wait_object_0, wait_abandoned):
            kernel32.ReleaseMutex(handle)
        kernel32.CloseHandle(handle)


def _next_data_row(layout: SheetLayout) -> int:
    key_column = layout.columns["材料名称"]
    last_row = layout.header_row
    for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
        if _text(layout.sheet.cell(row, key_column).value):
            last_row = row
    return last_row + 1


def _max_code_number(workbook: Any, layout: SheetLayout | None = None) -> int:
    maximum = 0
    targets = [layout] if layout is not None else [
        _locate_project_layout(workbook, sheet.title)
        for sheet in list_project_sheets(workbook)
    ]
    for target in targets:
        code_column = target.columns.get("物料编码")
        if not code_column:
            continue
        for row in range(target.header_row + 1, target.sheet.max_row + 1):
            match = MATERIAL_CODE_PATTERN.match(
                _text(target.sheet.cell(row, code_column).value)
            )
            if match:
                maximum = max(maximum, int(match.group(1)))
    if LEDGER_SHEET_NAME in workbook.sheetnames:
        ledger = workbook[LEDGER_SHEET_NAME]
        for row in range(2, ledger.max_row + 1):
            match = MATERIAL_CODE_PATTERN.match(_text(ledger.cell(row, 2).value))
            if match:
                maximum = max(maximum, int(match.group(1)))
    return maximum


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _column_width(sheet: Any, column: int, default: float) -> float:
    width = sheet.column_dimensions[get_column_letter(column)].width
    return float(width) if width is not None else default


def _insert_location_column(workbook: Any, layout: SheetLayout) -> SheetLayout:
    sheet = layout.sheet
    location_column = 9
    detail_width = _column_width(sheet, 9, 58)
    note_width = _column_width(sheet, 10, 48)
    sheet.insert_cols(location_column, 1)
    for row in range(
        layout.header_row,
        max(layout.header_row + 1, sheet.max_row) + 1,
    ):
        source = sheet.cell(row, 11)
        target = sheet.cell(row, location_column)
        _copy_cell_style(source, target)
        target.value = LOCATION_HEADER if row == layout.header_row else None
    sheet.column_dimensions["I"].width = 18
    sheet.column_dimensions["J"].width = detail_width
    sheet.column_dimensions["K"].width = note_width
    upgraded = SheetLayout(
        sheet,
        layout.header_row,
        {
            name: index + 1
            for index, name in enumerate(LEGACY_LOCATION_HEADERS)
        },
        legacy_stock_order=True,
    )
    _rebuild_main_table(
        workbook,
        sheet,
        header_row=upgraded.header_row,
        last_row=max(upgraded.header_row + 1, sheet.max_row),
    )
    return upgraded


def _move_appended_location_column(
    workbook: Any,
    layout: SheetLayout,
) -> SheetLayout:
    sheet = layout.sheet
    detail_width = _column_width(sheet, 9, 58)
    note_width = _column_width(sheet, 10, 48)
    location_width = _column_width(sheet, 11, 18)
    sheet.insert_cols(9, 1)
    for row in range(
        layout.header_row,
        max(layout.header_row + 1, sheet.max_row) + 1,
    ):
        source = sheet.cell(row, 12)
        target = sheet.cell(row, 9)
        _copy_cell_style(source, target)
        target.value = source.value
    sheet.delete_cols(12, 1)
    sheet.column_dimensions["I"].width = location_width
    sheet.column_dimensions["J"].width = detail_width
    sheet.column_dimensions["K"].width = note_width
    upgraded = SheetLayout(
        sheet,
        layout.header_row,
        {
            name: index + 1
            for index, name in enumerate(LEGACY_LOCATION_HEADERS)
        },
        legacy_stock_order=True,
    )
    _rebuild_main_table(
        workbook,
        sheet,
        header_row=upgraded.header_row,
        last_row=max(upgraded.header_row + 1, sheet.max_row),
    )
    return upgraded


def _move_stock_after_specification(
    workbook: Any,
    layout: SheetLayout,
) -> SheetLayout:
    sheet = layout.sheet
    original_widths = {
        column: _column_width(sheet, column, 13)
        for column in range(1, len(LEGACY_LOCATION_HEADERS) + 1)
    }
    sheet.insert_cols(4, 1)
    for row in range(
        layout.header_row,
        max(layout.header_row + 1, sheet.max_row) + 1,
    ):
        source = sheet.cell(row, 9)
        target = sheet.cell(row, 4)
        _copy_cell_style(source, target)
        target.value = source.value
        target.comment = copy(source.comment)
        if source.hyperlink:
            target._hyperlink = copy(source.hyperlink)
    sheet.delete_cols(9, 1)

    source_columns = (1, 2, 3, 8, 4, 5, 6, 7, 9, 10, 11)
    for target_column, source_column in enumerate(source_columns, start=1):
        sheet.column_dimensions[get_column_letter(target_column)].width = (
            original_widths[source_column]
        )

    upgraded = SheetLayout(
        sheet,
        layout.header_row,
        {name: index + 1 for index, name in enumerate(PREVIOUS_HEADERS)},
        location_after_date=True,
    )
    _rebuild_main_table(
        workbook,
        sheet,
        header_row=upgraded.header_row,
        last_row=max(upgraded.header_row + 1, sheet.max_row),
    )
    return upgraded


def _move_location_after_stock(
    workbook: Any,
    layout: SheetLayout,
) -> SheetLayout:
    sheet = layout.sheet
    original_widths = {
        column: _column_width(sheet, column, 13)
        for column in range(1, len(PREVIOUS_HEADERS) + 1)
    }
    sheet.insert_cols(5, 1)
    for row in range(
        layout.header_row,
        max(layout.header_row + 1, sheet.max_row) + 1,
    ):
        source = sheet.cell(row, 10)
        target = sheet.cell(row, 5)
        _copy_cell_style(source, target)
        target.value = source.value
        target.comment = copy(source.comment)
        if source.hyperlink:
            target._hyperlink = copy(source.hyperlink)
    sheet.delete_cols(10, 1)

    source_columns = (1, 2, 3, 4, 9, 5, 6, 7, 8, 10, 11)
    for target_column, source_column in enumerate(source_columns, start=1):
        sheet.column_dimensions[get_column_letter(target_column)].width = (
            original_widths[source_column]
        )

    upgraded = SheetLayout(
        sheet,
        layout.header_row,
        {name: index + 1 for index, name in enumerate(HEADERS)},
    )
    _rebuild_main_table(
        workbook,
        sheet,
        header_row=upgraded.header_row,
        last_row=max(upgraded.header_row + 1, sheet.max_row),
    )
    return upgraded


def _upgrade_main_sheet(workbook: Any, layout: SheetLayout) -> SheetLayout:
    if layout.temp_v2:
        sheet = layout.sheet
        sheet.insert_cols(1, 1)
        sheet.insert_cols(3, 1)
        for column, header in enumerate(PRE_LOCATION_HEADERS, start=1):
            sheet.cell(layout.header_row, column).value = header

        header_style_source = sheet.cell(layout.header_row, 2)
        for column in (1, 3):
            target = sheet.cell(layout.header_row, column)
            if header_style_source.has_style:
                target._style = copy(header_style_source._style)
            target.alignment = copy(header_style_source.alignment)
            target.number_format = header_style_source.number_format

        layout = SheetLayout(
            sheet,
            layout.header_row,
            {
                name: index + 1
                for index, name in enumerate(PRE_LOCATION_HEADERS)
            },
            pre_location=True,
        )
        next_number = _max_code_number(workbook, layout)
        for row in range(layout.header_row + 1, sheet.max_row + 1):
            if not _text(sheet.cell(row, layout.columns["材料名称"]).value):
                continue
            for target_column, source_column in ((1, 2), (3, 2)):
                source = sheet.cell(row, source_column)
                target = sheet.cell(row, target_column)
                if source.has_style:
                    target._style = copy(source._style)
                target.alignment = copy(source.alignment)
                target.number_format = source.number_format
            if not _text(sheet.cell(row, layout.columns["物料编码"]).value):
                next_number += 1
                sheet.cell(row, layout.columns["物料编码"]).value = (
                    f"MAT-{next_number:06d}"
                )
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["C"].width = 26

    if layout.appended_location:
        layout = _move_appended_location_column(workbook, layout)
    elif layout.pre_location:
        layout = _insert_location_column(workbook, layout)
    if layout.legacy_stock_order:
        layout = _move_stock_after_specification(workbook, layout)
    if layout.location_after_date:
        layout = _move_location_after_stock(workbook, layout)
    return layout


def _apply_date_only_formats(workbook: Any, layout: SheetLayout) -> bool:
    changed = False
    date_column = layout.columns["入库时间"]
    for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
        cell = layout.sheet.cell(row, date_column)
        if cell.value in (None, ""):
            continue
        if isinstance(cell.value, datetime) and any(
            (cell.value.hour, cell.value.minute, cell.value.second, cell.value.microsecond)
        ):
            cell.value = cell.value.replace(
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )
            changed = True
        if cell.number_format != "yyyy-mm-dd":
            cell.number_format = "yyyy-mm-dd"
            changed = True

    if LEDGER_SHEET_NAME in workbook.sheetnames:
        ledger = workbook[LEDGER_SHEET_NAME]
        for row in range(2, ledger.max_row + 1):
            cell = ledger.cell(row, 8)
            if cell.value in (None, ""):
                continue
            if isinstance(cell.value, datetime) and any(
                (
                    cell.value.hour,
                    cell.value.minute,
                    cell.value.second,
                    cell.value.microsecond,
                )
            ):
                cell.value = cell.value.replace(
                    hour=0,
                    minute=0,
                    second=0,
                    microsecond=0,
                )
                changed = True
            if cell.number_format != "yyyy-mm-dd":
                cell.number_format = "yyyy-mm-dd"
                changed = True
    return changed


def _apply_requester_normalization(workbook: Any, layout: SheetLayout) -> bool:
    changed = False
    requester_column = layout.columns["需求者"]
    for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
        cell = layout.sheet.cell(row, requester_column)
        replacement = REQUESTER_REPLACEMENTS.get(_text(cell.value))
        if replacement is not None and cell.value != replacement:
            cell.value = replacement
            changed = True

    if LEDGER_SHEET_NAME in workbook.sheetnames:
        ledger = workbook[LEDGER_SHEET_NAME]
        for row in range(2, ledger.max_row + 1):
            cell = ledger.cell(row, 6)
            replacement = REQUESTER_REPLACEMENTS.get(_text(cell.value))
            if replacement is not None and cell.value != replacement:
                cell.value = replacement
                changed = True
    return changed


def _is_standard_stock_rule(rule: Any) -> bool:
    return (
        getattr(rule, "type", None) == "cellIs"
        and getattr(rule, "operator", None) in {"equal", "greaterThan"}
        and [str(value) for value in (getattr(rule, "formula", None) or [])]
        == ["0"]
    )


def _stock_rule_color(rule: Any) -> str:
    dxf = getattr(rule, "dxf", None)
    font = getattr(dxf, "font", None) if dxf is not None else None
    color = getattr(font, "color", None) if font is not None else None
    return str(getattr(color, "rgb", "") or "").upper()


def _normalize_stock_conditional_formatting(layout: SheetLayout) -> bool:
    sheet = layout.sheet
    stock_letter = get_column_letter(layout.columns["库存数量"])
    first_row = layout.header_row + 1
    last_row = max(first_row, sheet.max_row)
    target_range = (
        f"{stock_letter}{first_row}"
        if first_row == last_row
        else f"{stock_letter}{first_row}:{stock_letter}{last_row}"
    )
    expected_colors = {
        "equal": "FF9C0006",
        "greaterThan": "FF375623",
    }
    matching = [
        (conditional_format, rule)
        for conditional_format, rules in sheet.conditional_formatting._cf_rules.items()
        for rule in rules
        if _is_standard_stock_rule(rule)
    ]
    if (
        len(matching) == 2
        and {rule.operator for _, rule in matching} == set(expected_colors)
        and all(str(conditional_format.sqref) == target_range for conditional_format, _ in matching)
        and all(
            _stock_rule_color(rule) == expected_colors[rule.operator]
            for _, rule in matching
        )
    ):
        return False

    for conditional_format in list(sheet.conditional_formatting._cf_rules):
        retained = [
            rule
            for rule in sheet.conditional_formatting._cf_rules[conditional_format]
            if not _is_standard_stock_rule(rule)
        ]
        if retained:
            sheet.conditional_formatting._cf_rules[conditional_format] = retained
        else:
            del sheet.conditional_formatting._cf_rules[conditional_format]

    sheet.conditional_formatting.add(
        target_range,
        CellIsRule(
            operator="equal",
            formula=["0"],
            font=Font(color=expected_colors["equal"]),
        ),
    )
    sheet.conditional_formatting.add(
        target_range,
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            font=Font(color=expected_colors["greaterThan"]),
        ),
    )
    return True


def _apply_center_alignment(
    workbook: Any,
    main_layout: SheetLayout | None = None,
) -> bool:
    changed = False
    detail_index = HEADERS.index("变更明细") + 1
    for sheet in workbook.worksheets:
        table_ranges = [
            range_boundaries(table.ref)
            for table in sheet.tables.values()
        ]
        if not table_ranges:
            table_ranges = [(1, 1, sheet.max_column, sheet.max_row)]
        for min_column, min_row, max_column, max_row in table_ranges:
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_column,
                max_col=max_column,
            ):
                for cell in row:
                    is_main_detail = (
                        is_project_sheet_title(sheet.title)
                        and cell.column == detail_index
                        and cell.row > 1
                    )
                    horizontal = "left" if is_main_detail else "center"
                    wrap_text = True if is_main_detail else cell.alignment.wrap_text
                    if (
                        cell.alignment.horizontal == horizontal
                        and cell.alignment.vertical == "center"
                        and cell.alignment.wrap_text == wrap_text
                    ):
                        continue
                    alignment = copy(cell.alignment)
                    alignment.horizontal = horizontal
                    alignment.vertical = "center"
                    alignment.wrap_text = wrap_text
                    cell.alignment = alignment
                    changed = True
    return changed


def _apply_main_column_widths(layout: SheetLayout) -> bool:
    changed = False
    for header, expected_width in MAIN_COLUMN_WIDTHS.items():
        column_letter = get_column_letter(layout.columns[header])
        dimension = layout.sheet.column_dimensions[column_letter]
        current_width = dimension.width
        if current_width is not None and math.isclose(
            float(current_width),
            expected_width,
            rel_tol=0,
            abs_tol=0.01,
        ):
            continue
        dimension.width = expected_width
        changed = True
    return changed


def _merge_pending_sheet(workbook: Any) -> bool:
    """把遗留的「待确认物料」表按项目合并进对应项目 sheet，然后删除该表。"""
    if PENDING_SHEET_NAME not in workbook.sheetnames:
        return False
    pending = workbook[PENDING_SHEET_NAME]
    headers: dict[str, int] = {}
    for column in range(1, pending.max_column + 1):
        headers[_text(pending.cell(1, column).value)] = column

    def hv(name: str) -> int | None:
        return headers.get(name)

    rows: list[tuple[str, str, str, Any, str, str, str, str]] = []
    for row in range(2, pending.max_row + 1):
        project = _text(pending.cell(row, hv("项目")).value if hv("项目") else None)
        if not project:
            continue
        material = _text(pending.cell(row, hv("Material Name")).value if hv("Material Name") else "")
        desc = _text(pending.cell(row, hv("Description")).value if hv("Description") else "")
        series = _text(pending.cell(row, hv("Series")).value if hv("Series") else "")
        requestor = _text(pending.cell(row, hv("Requestor")).value if hv("Requestor") else "")
        po = _text(pending.cell(row, hv("PO")).value if hv("PO") else "")
        receive_time = _text(pending.cell(row, hv("收货时间")).value if hv("收货时间") else "")
        receive_qty = pending.cell(row, hv("收到数量")).value if hv("收到数量") else None
        remain_qty = pending.cell(row, hv("剩余数量")).value if hv("剩余数量") else None
        brand = _text(pending.cell(row, hv("Brand")).value if hv("Brand") else "")
        reason = _text(pending.cell(row, hv("待确认原因")).value if hv("待确认原因") else "")
        note_parts = ["原待确认物料"]
        if brand:
            note_parts.append(f"Brand:{brand}")
        if material:
            note_parts.append(f"MaterialName:{material}")
        if desc and desc != material:
            note_parts.append(f"Description:{desc}")
        if receive_time:
            note_parts.append(f"收货时间:{receive_time}")
        if receive_qty not in (None, ""):
            note_parts.append(f"收到数量:{receive_qty}")
        if remain_qty not in (None, ""):
            note_parts.append(f"剩余数量:{remain_qty}")
        if reason:
            note_parts.append(f"待确认原因:{reason}")
        rows.append(
            (
                project,
                material or desc,
                series,
                remain_qty,
                requestor,
                po,
                receive_time,
                "；".join(note_parts),
            )
        )
    for project, material, series, remain, requestor, po, rtime, note in rows:
        sheet = find_project_sheet(workbook, project)
        if sheet is None:
            sheet = _create_project_sheet(workbook, project)
        target_row = sheet.max_row + 1
        if target_row > 2:
            _copy_row_style(sheet, 2, target_row, len(HEADERS))
        stock = None
        if remain not in (None, ""):
            if isinstance(remain, (int, float)) and not isinstance(remain, bool):
                stock = remain
            else:
                try:
                    stock = float(str(remain).replace(",", "").strip())
                except (TypeError, ValueError):
                    stock = None
        values = {
            "物料编码": None,
            "材料名称": material,
            "规格型号": series,
            "库存数量": stock,
            "物料放置位置": None,
            "项目": project,
            "需求者": requestor,
            "PO": po,
            "入库时间": rtime,
            "变更明细": None,
            "备注": note,
        }
        for index, name in enumerate(HEADERS, start=1):
            cell = sheet.cell(target_row, index)
            cell.value = values.get(name)
            if name == "库存数量" and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0;[Red]-#,##0"
    del workbook[PENDING_SHEET_NAME]
    return True


def _polish_workbook(workbook: Any) -> bool:
    """表格观感优化：冻结首行、表头行高，让滚动查看更舒适。"""
    changed = False
    for sheet in workbook.worksheets:
        if sheet.max_row >= 1:
            if sheet.freeze_panes != "A2":
                sheet.freeze_panes = "A2"
                changed = True
            current_height = sheet.row_dimensions[1].height
            if current_height != 22:
                sheet.row_dimensions[1].height = 22
                changed = True
    return changed


def upgrade_workbook_schema(path: Path) -> bool:
    if path.suffix.lower() != ".xlsx":
        raise ValueError("请选择 .xlsx 格式的 Excel 文件。")
    if not path.is_file():
        raise FileNotFoundError(f"找不到 Excel 文件：{path}")

    with WRITE_LOCK, _cross_process_workbook_lock(path):
        workbook = load_workbook(path, data_only=False)
        temp_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        try:
            changed = False
            changed = _split_single_sheet_to_projects(workbook) or changed
            changed = _merge_pending_sheet(workbook) or changed
            changed = _polish_workbook(workbook) or changed
            layouts = [
                _locate_project_layout(workbook, sheet.title)
                for sheet in list_project_sheets(workbook)
            ]
            if not layouts:
                _create_project_sheet(workbook, "未分类")
                layouts = [_locate_project_layout(workbook, "未分类")]
                changed = True
            for layout in layouts:
                layout_upgrade_needed = any(
                    (
                        layout.temp_v2,
                        layout.pre_location,
                        layout.appended_location,
                        layout.legacy_stock_order,
                        layout.location_after_date,
                    )
                )
                if layout_upgrade_needed:
                    layout = _upgrade_main_sheet(workbook, layout)
                changed = (
                    _apply_date_only_formats(workbook, layout)
                    or changed
                )
                changed = (
                    _apply_requester_normalization(workbook, layout)
                    or changed
                )
                changed = (
                    _normalize_stock_conditional_formatting(layout)
                    or changed
                )
                changed = _apply_center_alignment(workbook, layout) or changed
                changed = _apply_main_column_widths(layout) or changed
            if not changed:
                return False
            if LEDGER_SHEET_NAME in workbook.sheetnames:
                workbook[LEDGER_SHEET_NAME].sheet_state = "hidden"
            calculation = getattr(workbook, "calculation", None)
            if calculation is not None:
                calculation.calcMode = "auto"
                calculation.fullCalcOnLoad = True
                calculation.forceFullCalc = True
            create_backup(path)
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, path)
            return True
        except PermissionError as exc:
            raise PermissionError(
                "Excel 文件正在被占用。请关闭桌面版 Excel 后再升级数据表。"
            ) from exc
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    pass


def _find_material_rows(
    layout: SheetLayout,
    *,
    material_code: str = "",
    material: str = "",
    specification: str = "",
) -> list[int]:
    code_target = _text(material_code).casefold()
    name_target = normalized_material(material)
    spec_target = normalized_specification(specification)
    rows = []
    for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
        if not _text(layout.sheet.cell(row, layout.columns["材料名称"]).value):
            continue
        if code_target:
            code_column = layout.columns.get("物料编码")
            if code_column and _text(layout.sheet.cell(row, code_column).value).casefold() == code_target:
                rows.append(row)
            continue
        row_name = normalized_material(layout.sheet.cell(row, layout.columns["材料名称"]).value)
        spec_column = layout.columns.get("规格型号")
        row_spec = normalized_specification(
            layout.sheet.cell(row, spec_column).value if spec_column else ""
        )
        if row_name == name_target and row_spec == spec_target:
            rows.append(row)
    return rows


def current_balance(
    path: Path,
    material: str = "",
    material_code: str = "",
    specification: str = "",
) -> float:
    if not _text(material_code) and not _text(material):
        return 0.0
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        matches = []
        for layout in [
            _locate_project_layout(workbook, sheet.title)
            for sheet in list_project_sheets(workbook)
        ]:
            rows = _find_material_rows(
                layout,
                material_code=material_code,
                material=material,
                specification=specification,
            )
            for row in rows:
                matches.append((layout, row))
        if len(matches) > 1:
            raise ValueError("找到多个同名同规格物料，请通过物料编码选择。")
        if not matches:
            return 0.0
        layout, row = matches[0]
        return _stock_value(
            layout.sheet.cell(row, layout.columns["库存数量"]).value
        )
    finally:
        workbook.close()


def _ensure_ledger(workbook: Any) -> Any:
    if LEDGER_SHEET_NAME in workbook.sheetnames:
        ledger = workbook[LEDGER_SHEET_NAME]
        if _row_values(ledger, 1, len(LEDGER_HEADERS)) != LEDGER_HEADERS:
            raise ValueError(f"隐藏工作表“{LEDGER_SHEET_NAME}”表头不正确。")
    else:
        ledger = workbook.create_sheet(LEDGER_SHEET_NAME)
        for column, header in enumerate(LEDGER_HEADERS, start=1):
            ledger.cell(1, column).value = header
        for column in range(1, len(LEDGER_HEADERS) + 1):
            ledger.cell(2, column).value = None
        table = Table(
            displayName=_next_table_name(workbook, "InventoryLedgerTable"),
            ref="A1:M2",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ledger.add_table(table)
    ledger.sheet_state = "hidden"
    return ledger


def _next_ledger_code(ledger: Any) -> str:
    maximum = 0
    for row in range(2, ledger.max_row + 1):
        match = LEDGER_CODE_PATTERN.match(_text(ledger.cell(row, 1).value))
        if match:
            maximum = max(maximum, int(match.group(1)))
    return f"TXN-{maximum + 1:08d}"


def _detail_line(record: InventoryRecord) -> str:
    display_quantity = (
        abs(record.quantity)
        if record.change_type in POSITIVE_TYPES | NEGATIVE_TYPES
        else record.quantity
    )
    return (
        f"{record.changed_at.year}-{record.changed_at.month}-{record.changed_at.day}："
        f"{_inline_text(record.requester)}"
        f"{_inline_text(record.change_type)}"
        f"{format_number(display_quantity)}个；"
    )


def _inline_text(value: Any) -> str:
    return " ".join(_text(value).replace("；", " ").replace(";", " ").split())


def _append_detail(existing: Any, line: str) -> str:
    previous = _text(existing).rstrip("；; \r\n")
    return f"{previous}；\n{line}" if previous else line


def _estimated_detail_lines(value: Any, column_width: float | None) -> int:
    capacity = max(12, int(column_width or 58))
    visual_lines = 0
    for line in str(value or "").splitlines() or [""]:
        display_units = sum(
            2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
            for character in line
        )
        visual_lines += max(1, math.ceil(display_units / capacity))
    return visual_lines


def _next_ledger_row(ledger: Any) -> int:
    for row in range(2, ledger.max_row + 1):
        if not any(
            _text(ledger.cell(row, column).value)
            for column in range(1, len(LEDGER_HEADERS) + 1)
        ):
            return row
    return max(2, ledger.max_row + 1)


def create_backup(path: Path) -> None:
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = backup_dir / f"{path.stem}_{stamp}.xlsx"
    shutil.copy2(path, backup_path)
    backups = sorted(
        backup_dir.glob(f"{path.stem}_*.xlsx"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for old_backup in backups[MAX_BACKUPS:]:
        try:
            old_backup.unlink()
        except OSError:
            pass


def write_record(path: Path, record: InventoryRecord) -> tuple[float, float, int]:
    with WRITE_LOCK, _cross_process_workbook_lock(path):
        validate_record(
            record.material,
            record.requester,
            record.project,
            record.change_type,
            str(record.quantity),
            record.specification,
            record.material_code,
            record.location,
        )
        _validate_free_text(record.po, "PO", required=False, limit=200)
        _validate_free_text(record.note, "备注", required=False, limit=2000)
        if not isinstance(record.changed_at, datetime):
            raise ValueError("变动时间格式不正确。")
        record.changed_at = record.changed_at.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
        )
        validate_workbook(path)
        workbook = load_workbook(path, data_only=False)
        temp_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        try:
            layout = _locate_project_layout(workbook, record.project)
            ledger = _ensure_ledger(workbook)
            _apply_date_only_formats(workbook, layout)
            _apply_requester_normalization(workbook, layout)
            rows = _find_material_rows(
                layout,
                material_code=record.material_code,
                material=record.material,
                specification=record.specification,
            )
            if record.material_code and not rows:
                # 物料编码全局唯一：跨所有项目 sheet 定位物料
                for other_layout in [
                    _locate_project_layout(workbook, sheet.title)
                    for sheet in list_project_sheets(workbook)
                ]:
                    if other_layout.sheet is layout.sheet:
                        continue
                    candidate_rows = _find_material_rows(
                        other_layout,
                        material_code=record.material_code,
                    )
                    if candidate_rows:
                        layout = other_layout
                        rows = candidate_rows
                        record.project = other_layout.sheet.title
                        break
            if record.material_code and not rows:
                raise ValueError(f"找不到物料编码“{record.material_code}”。")
            if len(rows) > 1:
                raise ValueError("找到多个同名同规格物料，请通过物料编码选择。")

            is_new = not rows
            if is_new:
                # 新物料：跨所有项目 sheet 查重，防止同一物料散落在多个项目
                for other_layout in [
                    _locate_project_layout(workbook, sheet.title)
                    for sheet in list_project_sheets(workbook)
                ]:
                    if other_layout.sheet is layout.sheet:
                        continue
                    duplicates = _find_material_rows(
                        other_layout,
                        material=record.material,
                        specification=record.specification,
                    )
                    if duplicates:
                        raise ValueError(
                            f"该物料已存在于项目“{other_layout.sheet.title}”，"
                            "请在对应项目下选择该物料。"
                        )
                row = _next_data_row(layout)
                _copy_row_style(
                    layout.sheet,
                    max(layout.header_row + 1, row - 1),
                    row,
                    len(HEADERS),
                )
                next_code = _max_code_number(workbook) + 1
                record.material_code = f"MAT-{next_code:06d}"
                balance_before = 0.0
            else:
                row = rows[0]
                code_column = layout.columns["物料编码"]
                record.material_code = _text(layout.sheet.cell(row, code_column).value)
                if not record.material_code:
                    next_code = _max_code_number(workbook) + 1
                    record.material_code = f"MAT-{next_code:06d}"
                    layout.sheet.cell(row, code_column).value = record.material_code
                stored_material = _text(
                    layout.sheet.cell(row, layout.columns["材料名称"]).value
                )
                stored_specification = _text(
                    layout.sheet.cell(row, layout.columns["规格型号"]).value
                )
                if (
                    normalized_material(stored_material)
                    != normalized_material(record.material)
                    or normalized_specification(stored_specification)
                    != normalized_specification(record.specification)
                ):
                    raise ValueError(
                        "所选物料的名称或规格已发生变化，请重新搜索并选择物料。"
                    )
                balance_before = _stock_value(
                    layout.sheet.cell(row, layout.columns["库存数量"]).value
                )
                if not record.location:
                    record.location = _text(
                        layout.sheet.cell(
                            row,
                            layout.columns[LOCATION_HEADER],
                        ).value
                    )

                duplicate_rows = _find_material_rows(
                    layout,
                    material=record.material,
                    specification=record.specification,
                )
                if any(candidate != row for candidate in duplicate_rows):
                    raise ValueError("修改后的材料名称和规格型号与另一物料重复。")

            balance_after = balance_before + record.quantity
            if balance_after < -1e-9:
                raise ValueError(
                    f"库存不足：当前为 {format_number(balance_before)}，"
                    f"本次变动后将为 {format_number(balance_after)}。"
                )

            sheet = layout.sheet
            values = {
                "物料编码": record.material_code,
                "项目": record.project,
                "需求者": record.requester,
                "PO": record.po,
                "库存数量": balance_after,
            }
            if is_new:
                values["材料名称"] = record.material
                values["规格型号"] = record.specification
                values["入库时间"] = record.changed_at
                values[LOCATION_HEADER] = record.location
            elif record.location:
                values[LOCATION_HEADER] = record.location
            for field, value in values.items():
                sheet.cell(row, layout.columns[field]).value = value
            if record.note or is_new:
                sheet.cell(row, layout.columns["备注"]).value = record.note
            _extend_header_table(
                sheet,
                header_row=layout.header_row,
                last_row=row,
                last_column=len(HEADERS),
            )

            detail_cell = sheet.cell(row, layout.columns["变更明细"])
            detail_cell.value = _append_detail(detail_cell.value, _detail_line(record))
            detail_cell.alignment = copy(detail_cell.alignment)
            detail_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                text_rotation=detail_cell.alignment.text_rotation,
                wrap_text=True,
                shrink_to_fit=detail_cell.alignment.shrink_to_fit,
                indent=detail_cell.alignment.indent,
            )
            detail_column_letter = get_column_letter(layout.columns["变更明细"])
            detail_lines = _estimated_detail_lines(
                detail_cell.value,
                sheet.column_dimensions[detail_column_letter].width,
            )
            sheet.row_dimensions[row].height = min(
                409,
                max(24, 18 * detail_lines),
            )
            sheet.cell(row, layout.columns["入库时间"]).number_format = "yyyy-mm-dd"
            sheet.cell(row, layout.columns["库存数量"]).number_format = (
                "#,##0;[Red]-#,##0"
            )

            ledger_row = _next_ledger_row(ledger)
            if ledger_row > 2:
                _copy_row_style(ledger, 2, ledger_row, len(LEDGER_HEADERS))
            ledger_values = [
                _next_ledger_code(ledger),
                record.material_code,
                record.material,
                record.specification,
                record.project,
                record.requester,
                record.po,
                record.changed_at,
                record.change_type,
                record.quantity,
                balance_before,
                balance_after,
                record.note,
            ]
            for column, value in enumerate(ledger_values, start=1):
                ledger.cell(ledger_row, column).value = value
            ledger.cell(ledger_row, 8).number_format = "yyyy-mm-dd"
            for column in (10, 11, 12):
                ledger.cell(ledger_row, column).number_format = (
                    "#,##0;[Red]-#,##0"
                )
            _extend_header_table(
                ledger,
                header_row=1,
                last_row=ledger_row,
                last_column=len(LEDGER_HEADERS),
            )
            _normalize_stock_conditional_formatting(layout)
            _apply_center_alignment(workbook, layout)
            _apply_main_column_widths(layout)

            calculation = getattr(workbook, "calculation", None)
            if calculation is not None:
                calculation.calcMode = "auto"
                calculation.fullCalcOnLoad = True
                calculation.forceFullCalc = True

            create_backup(path)
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, path)
            return balance_before, balance_after, row
        except PermissionError as exc:
            raise PermissionError(
                "Excel 文件正在被占用。请关闭桌面版 Excel 后再提交。"
            ) from exc
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    pass


def recent_records(path: Path, limit: int = 8) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if LEDGER_SHEET_NAME not in workbook.sheetnames:
            return []
        ledger = workbook[LEDGER_SHEET_NAME]
        if _row_values(ledger, 1, len(LEDGER_HEADERS)) != LEDGER_HEADERS:
            raise ValueError(f"隐藏工作表“{LEDGER_SHEET_NAME}”表头不正确。")
        locations_by_code = {}
        layouts = [
            _locate_project_layout(workbook, sheet.title)
            for sheet in list_project_sheets(workbook)
        ]
        for layout in layouts:
            location_column = layout.columns.get(LOCATION_HEADER)
            code_column = layout.columns.get("物料编码")
            if not location_column or not code_column:
                continue
            for material_row in range(
                layout.header_row + 1,
                layout.sheet.max_row + 1,
            ):
                material_code = _text(
                    layout.sheet.cell(material_row, code_column).value
                )
                if material_code:
                    locations_by_code[material_code.casefold()] = _text(
                        layout.sheet.cell(material_row, location_column).value
                    )
        records = []
        first_row = max(2, ledger.max_row - max(1, limit) + 1)
        for row in range(ledger.max_row, first_row - 1, -1):
            if not _text(ledger.cell(row, 1).value):
                continue
            changed_at = ledger.cell(row, 8).value
            records.append(
                {
                    "transaction_id": _text(ledger.cell(row, 1).value),
                    "material_code": _text(ledger.cell(row, 2).value),
                    "material": _text(ledger.cell(row, 3).value),
                    "specification": _text(ledger.cell(row, 4).value),
                    "project": _text(ledger.cell(row, 5).value),
                    "requester": _text(ledger.cell(row, 6).value),
                    "location": locations_by_code.get(
                        _text(ledger.cell(row, 2).value).casefold(),
                        "",
                    ),
                    "changed_at": (
                        changed_at.strftime("%Y-%m-%d")
                        if isinstance(changed_at, datetime)
                        else _text(changed_at)
                    ),
                    "change_type": _text(ledger.cell(row, 9).value),
                    "quantity": numeric_value(ledger.cell(row, 10).value),
                    "balance": numeric_value(ledger.cell(row, 12).value),
                }
            )
            if len(records) >= limit:
                break
        return records
    finally:
        workbook.close()


def search_materials(path: Path, query: str = "", limit: int = 20) -> list[dict[str, Any]]:
    tokens = [
        _normalized_identity(token)
        for token in unicodedata.normalize("NFKC", query).split()
        if token.strip()
    ]
    safe_limit = max(1, min(int(limit), 100))
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        results = []
        layouts = [
            _locate_project_layout(workbook, sheet.title)
            for sheet in list_project_sheets(workbook)
        ]
        for layout in layouts:
            for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
                material = _text(
                    layout.sheet.cell(row, layout.columns["材料名称"]).value
                )
                if not material:
                    continue
                code_column = layout.columns.get("物料编码")
                spec_column = layout.columns.get("规格型号")
                code = _text(layout.sheet.cell(row, code_column).value) if code_column else ""
                specification = (
                    _text(layout.sheet.cell(row, spec_column).value) if spec_column else ""
                )
                location_column = layout.columns.get(LOCATION_HEADER)
                location = (
                    _text(layout.sheet.cell(row, location_column).value)
                    if location_column
                    else ""
                )
                haystack = _normalized_identity(
                    f"{code} {material} {specification} {location}"
                )
                if tokens and not all(token in haystack for token in tokens):
                    continue
                results.append(
                    {
                        "materialCode": code,
                        "material": material,
                        "specification": specification,
                        "project": _text(
                            layout.sheet.cell(row, layout.columns["项目"]).value
                        ) or layout.sheet.title,
                        "requester": _text(
                            layout.sheet.cell(row, layout.columns["需求者"]).value
                        ),
                        "po": _text(layout.sheet.cell(row, layout.columns["PO"]).value),
                        "location": location,
                        "balance": _stock_value(
                            layout.sheet.cell(row, layout.columns["库存数量"]).value
                        ),
                    }
                )
                if len(results) >= safe_limit:
                    break
            if len(results) >= safe_limit:
                break
        return results
    finally:
        workbook.close()


def workbook_suggestions(path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        suggestions = {"requesters": [], "projects": [], "locations": []}
        seen = {name: set() for name in suggestions}
        layouts = [
            _locate_project_layout(workbook, sheet.title)
            for sheet in list_project_sheets(workbook)
        ]
        for layout in layouts:
            for name, column in (
                ("requesters", layout.columns["需求者"]),
                ("locations", layout.columns.get(LOCATION_HEADER)),
            ):
                if not column:
                    continue
                for row in range(layout.header_row + 1, layout.sheet.max_row + 1):
                    value = _text(layout.sheet.cell(row, column).value)
                    if not value or value.lstrip().startswith("="):
                        continue
                    identity = _normalized_identity(value)
                    if identity in seen[name]:
                        continue
                    seen[name].add(identity)
                    suggestions[name].append(value)
            project_name = layout.sheet.title
            identity = _normalized_identity(project_name)
            if project_name and identity not in seen["projects"]:
                seen["projects"].add(identity)
                suggestions["projects"].append(project_name)
        return suggestions
    finally:
        workbook.close()


def format_number(value: float) -> str:
    return f"{value:,.0f}"


def parse_datetime(text: str) -> datetime:
    normalized = text.strip().replace("/", "-")
    try:
        return datetime.strptime(normalized, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(
            "变动时间请使用“年-月-日”，例如 2026-07-29。"
        ) from exc


def _validate_free_text(value: str, label: str, *, required: bool, limit: int) -> None:
    if required and not value.strip():
        raise ValueError(f"请填写{label}。")
    if len(value) > limit:
        raise ValueError(f"{label}不能超过 {limit} 个字符。")
    if any(ord(character) < 32 and character not in "\t" for character in value):
        raise ValueError(f"{label}不能包含换行或控制字符。")
    if value.lstrip().startswith("="):
        raise ValueError(f"{label}不能以等号开头。")


def validate_record(
    material: str,
    requester: str,
    project: str,
    change_type: str,
    quantity_text: str,
    specification: str = "",
    material_code: str = "",
    location: str = "",
) -> float:
    _validate_free_text(material, "材料名称", required=True, limit=200)
    _validate_free_text(specification, "规格型号", required=False, limit=200)
    _validate_free_text(requester, "需求者", required=True, limit=100)
    _validate_free_text(project, "项目", required=True, limit=100)
    _validate_free_text(material_code, "物料编码", required=False, limit=50)
    _validate_free_text(location, LOCATION_HEADER, required=False, limit=100)
    if material_code and not MATERIAL_CODE_PATTERN.match(material_code):
        raise ValueError("物料编码格式不正确，应类似 MAT-000001。")
    if change_type not in CHANGE_TYPES:
        raise ValueError("请选择有效的库存管理类型。")
    normalized_quantity = quantity_text.replace(",", "").strip()
    try:
        decimal_quantity = Decimal(normalized_quantity)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("变动数量必须是数字。") from exc
    if not decimal_quantity.is_finite():
        raise ValueError("变动数量必须是有限数字。")
    if abs(decimal_quantity) > MAX_ABS_QUANTITY:
        raise ValueError("变动数量过大，请填写不超过 1,000,000,000,000 的数值。")
    if decimal_quantity != decimal_quantity.to_integral_value():
        raise ValueError("变动数量必须是整数。")
    if decimal_quantity == 0:
        raise ValueError("变动数量不能为 0。")
    quantity = float(decimal_quantity)
    if change_type in POSITIVE_TYPES and quantity < 0:
        raise ValueError(f"“{change_type}”应填写正数。")
    if change_type in NEGATIVE_TYPES and quantity > 0:
        raise ValueError(f"“{change_type}”应填写负数。")
    return quantity


def run_self_test(path: Path) -> dict[str, Any]:
    validate_workbook(path)
    first = InventoryRecord(
        material="EXE测试材料",
        specification="TEST-01",
        requester="自定义需求者",
        po="PO-TEST-001",
        changed_at=datetime(2026, 7, 29, 10, 0),
        quantity=10,
        change_type="入库",
        project="自定义项目",
        note="自动化测试",
    )
    second = InventoryRecord(
        material="EXE测试材料",
        specification="TEST-01",
        requester="王五",
        po="",
        changed_at=datetime(2026, 7, 29, 11, 0),
        quantity=-3,
        change_type="领用",
        project="项目C",
        note="自动化测试",
    )
    _, first_after, first_row = write_record(path, first)
    second.material_code = first.material_code
    _, second_after, second_row = write_record(path, second)
    assert first_after == 10
    assert second_after == 7
    assert second_row == first_row
    assert current_balance(path, material_code=first.material_code) == 7
    assert search_materials(path, "EXE TEST")[0]["materialCode"] == first.material_code
    return {
        "status": "ok",
        "row": first_row,
        "materialCode": first.material_code,
        "balance": second_after,
    }


def create_project(path: Path, project: str) -> str:
    """管理员新增项目：创建工作簿中的项目 sheet。"""
    project = validate_project_name(project)
    with WRITE_LOCK, _cross_process_workbook_lock(path):
        workbook = load_workbook(path, data_only=False)
        temp_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        try:
            if find_project_sheet(workbook, project) is not None:
                raise ValueError(f"项目“{project}”已存在。")
            _create_project_sheet(workbook, project)
            create_backup(path)
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, path)
            return project
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    pass


def remove_project(path: Path, project: str) -> None:
    """管理员删除项目：删除对应项目 sheet（会丢失该项目数据，请先确认）。"""
    project = validate_project_name(project)
    with WRITE_LOCK, _cross_process_workbook_lock(path):
        workbook = load_workbook(path, data_only=False)
        temp_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        try:
            sheet = find_project_sheet(workbook, project)
            if sheet is None:
                raise ValueError(f"项目“{project}”不存在。")
            if len(list_project_sheets(workbook)) <= 1:
                raise ValueError("至少需要保留一个项目，不能删除最后一个。")
            del workbook[sheet.title]
            create_backup(path)
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, path)
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    pass
