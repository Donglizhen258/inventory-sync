"""测试旧单表模板迁移为多项目 sheet，并验证写入/搜索/项目CRUD。"""
import shutil
import tempfile
from pathlib import Path

import openpyxl

from inventory_core import (
    HEADERS,
    InventoryRecord,
    create_project,
    current_balance,
    list_project_names,
    remove_project,
    search_materials,
    upgrade_workbook_schema,
    write_record,
)

SRC = Path("assets/实验室库存管理模板.xlsx")
WORK = Path(tempfile.mkdtemp(prefix="erl_migrate_")) / "测试库存.xlsx"

print("=== 1. 复制旧模板并迁移 ===")
shutil.copy2(SRC, WORK)
changed = upgrade_workbook_schema(WORK)
print("迁移执行:", changed)

wb = openpyxl.load_workbook(WORK, data_only=False)
print("迁移后工作表:", wb.sheetnames)
projects = list_project_names(WORK)
print("项目列表:", projects)
assert len(projects) >= 2, "应有多个项目 sheet"
assert "库存管理" not in wb.sheetnames, "旧单表应被移除"
for name in projects:
    ws = wb[name]
    headers = [ws.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
    assert headers == HEADERS, f"{name} 表头错误: {headers}"

total_rows = sum(
    wb[name].max_row - 1 for name in projects if wb[name].max_row > 1
)
print("迁移后物料总行数:", total_rows)
assert total_rows >= 100, "物料数据应基本完整"

wb.close()
print("PASS 迁移")

print("\n=== 2. 按项目写入 ===")
record = InventoryRecord(
    material="迁移测试芯片",
    requester="王五",
    po="",
    changed_at=__import__("datetime").datetime(2026, 8, 12),
    quantity=10,
    change_type="入库",
    project=projects[0],
    note="迁移测试",
    specification="MIG-TEST-01",
    material_code="",
    location="货架A1",
)
before, after, row = write_record(WORK, record)
print(f"写入 {projects[0]} 成功: {before} -> {after} (row {row})")
assert record.material_code, "应自动生成物料编码"
assert current_balance(WORK, material_code=record.material_code) == 10
print("物料编码:", record.material_code)
print("PASS 写入")

print("\n=== 3. 跨项目搜索 ===")
results = search_materials(WORK, "迁移测试芯片", limit=5)
assert results and results[0]["material"] == "迁移测试芯片"
print("搜索到:", results[0]["materialCode"], results[0]["project"])
print("PASS 搜索")

print("\n=== 4. 新增项目 ===")
create_project(WORK, "新项目A")
projects2 = list_project_names(WORK)
assert "新项目A" in projects2
print("新增后项目:", projects2)
print("PASS 新增项目")

print("\n=== 5. 删除项目 ===")
remove_project(WORK, "新项目A")
projects3 = list_project_names(WORK)
assert "新项目A" not in projects3
print("删除后项目:", projects3)
print("PASS 删除项目")

print("\n=== 6. 项目重名校验 ===")
try:
    create_project(WORK, "_非法名")
    raise AssertionError("不应允许创建 _ 开头项目")
except ValueError as exc:
    print("拒绝非法项目名:", exc)
print("PASS 校验")

print("\n全部通过")
